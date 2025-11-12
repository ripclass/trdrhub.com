"""
Export Service
CSV/XLSX streaming exports for reporting
"""

import csv
import io
import logging
from typing import Generator, List, Dict, Any, Optional, Union
from datetime import datetime, date
from dataclasses import dataclass
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
import boto3
from sqlalchemy.orm import Session
from sqlalchemy import select, and_, or_, func, desc

from app.config import settings
from app.models.collaboration import Thread, Comment
from app.models.bulk_jobs import BulkJob, BulkJobItem
from app.models.workflows import WorkflowRule, PolicyOverride

logger = logging.getLogger(__name__)


@dataclass
class ExportFilter:
    """Export filter parameters"""
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    corridor: Optional[str] = None
    exporter: Optional[str] = None
    status: Optional[str] = None
    severity: Optional[str] = None
    tenant_alias: Optional[str] = None
    bank_alias: Optional[str] = None


class CSVExporter:
    """CSV export utility"""

    def __init__(self, headers: List[str], locale: str = "en"):
        self.headers = headers
        self.locale = locale

    def generate_csv(self, data_generator: Generator[List[Any], None, None]) -> Generator[str, None, None]:
        """Generate CSV content as string chunks"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        writer.writerow(self.headers)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        # Write data rows
        for row in data_generator:
            writer.writerow(row)
            content = output.getvalue()
            if content:  # Only yield if there's content
                yield content
                output.seek(0)
                output.truncate(0)


class XLSXExporter:
    """XLSX export utility"""

    def __init__(self, headers: List[str], locale: str = "en", sheet_name: str = "Export"):
        self.headers = headers
        self.locale = locale
        self.sheet_name = sheet_name
        self.temp_file = None

    def create_workbook(self, data_generator: Generator[List[Any], None, None]) -> bytes:
        """Create XLSX workbook in memory"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers with styling
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        row_num = 2
        for row_data in data_generator:
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
            row_num += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class ExportService:
    """Main export service"""

    def __init__(self, db: Session):
        self.db = db
        self.s3_client = None
        if settings.AWS_ACCESS_KEY_ID:
            self.s3_client = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_REGION
            )

    async def export_bank_portfolio(
        self,
        export_format: str,
        filters: ExportFilter,
        tenant_alias: str,
        user_locale: str = "en"
    ) -> StreamingResponse:
        """Export bank portfolio report"""
        headers = self._get_portfolio_headers(user_locale)

        def data_generator():
            # Query LCs with filters
            # Note: This would query actual LC data when implemented
            # For now, generating sample data based on collaboration threads and bulk jobs

            # Get threads as proxy for LC discussions
            query = select(Thread).where(Thread.tenant_alias == tenant_alias)

            if filters.start_date:
                query = query.where(Thread.created_at >= filters.start_date)
            if filters.end_date:
                query = query.where(Thread.created_at <= filters.end_date)
            if filters.status:
                query = query.where(Thread.status == filters.status)

            query = query.order_by(desc(Thread.created_at))

            result = self.db.execute(query)
            threads = result.scalars().all()

            for thread in threads:
                yield [
                    thread.id,
                    f"LC-{thread.id[:8].upper()}",  # Mock LC number
                    "USD-BD",  # Mock corridor
                    thread.created_by.get("name", "Unknown") if isinstance(thread.created_by, dict) else "Unknown",
                    thread.status.title(),
                    thread.priority.title(),
                    thread.created_at.strftime("%Y-%m-%d"),
                    len(thread.comments) if hasattr(thread, 'comments') else 0,
                    "Active" if thread.status == "open" else "Resolved"
                ]

        if export_format.lower() == "csv":
            exporter = CSVExporter(headers, user_locale)
            return StreamingResponse(
                exporter.generate_csv(data_generator()),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=bank_portfolio_{datetime.now().strftime('%Y%m%d')}.csv"}
            )
        elif export_format.lower() == "xlsx":
            return await self._handle_xlsx_export(
                headers, data_generator(), f"bank_portfolio_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
        else:
            raise HTTPException(status_code=400, detail="Unsupported export format")

    async def export_discrepancies_trend(
        self,
        export_format: str,
        filters: ExportFilter,
        tenant_alias: str,
        user_locale: str = "en"
    ) -> StreamingResponse:
        """Export discrepancies trend report"""
        headers = self._get_discrepancies_headers(user_locale)

        def data_generator():
            # Query high-priority threads as proxy for discrepancies
            query = select(Thread).where(
                and_(
                    Thread.tenant_alias == tenant_alias,
                    Thread.priority.in_(["high", "urgent"])
                )
            )

            if filters.start_date:
                query = query.where(Thread.created_at >= filters.start_date)
            if filters.end_date:
                query = query.where(Thread.created_at <= filters.end_date)

            query = query.order_by(desc(Thread.created_at))

            result = self.db.execute(query)
            threads = result.scalars().all()

            # Group by discrepancy type (mock based on thread title)
            discrepancy_counts = {}
            for thread in threads:
                # Mock discrepancy type extraction
                discrepancy_type = self._extract_discrepancy_type(thread.title)
                period = thread.created_at.strftime("%Y-%m-%d")
                key = (discrepancy_type, period)

                if key not in discrepancy_counts:
                    discrepancy_counts[key] = 0
                discrepancy_counts[key] += 1

            for (discrepancy_type, period), count in discrepancy_counts.items():
                yield [
                    period,
                    discrepancy_type,
                    count,
                    "High" if count > 5 else "Normal",
                    f"{count * 2.5:.1f}%"  # Mock resolution rate
                ]

        if export_format.lower() == "csv":
            exporter = CSVExporter(headers, user_locale)
            return StreamingResponse(
                exporter.generate_csv(data_generator()),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=discrepancies_trend_{datetime.now().strftime('%Y%m%d')}.csv"}
            )
        elif export_format.lower() == "xlsx":
            return await self._handle_xlsx_export(
                headers, data_generator(), f"discrepancies_trend_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
        else:
            raise HTTPException(status_code=400, detail="Unsupported export format")

    async def export_turnaround_report(
        self,
        export_format: str,
        filters: ExportFilter,
        tenant_alias: str,
        user_locale: str = "en"
    ) -> StreamingResponse:
        """Export turnaround time report"""
        headers = self._get_turnaround_headers(user_locale)

        def data_generator():
            # Query resolved threads to calculate turnaround times
            query = select(Thread).where(
                and_(
                    Thread.tenant_alias == tenant_alias,
                    Thread.status == "resolved"
                )
            )

            if filters.start_date:
                query = query.where(Thread.created_at >= filters.start_date)
            if filters.end_date:
                query = query.where(Thread.created_at <= filters.end_date)

            result = self.db.execute(query)
            threads = result.scalars().all()

            # Calculate turnaround statistics
            for thread in threads:
                # Mock resolution time calculation
                resolution_time = self._calculate_resolution_time(thread)

                yield [
                    f"EXP-{thread.id[:6].upper()}",  # Mock exporter ID
                    thread.created_by.get("name", "Unknown") if isinstance(thread.created_by, dict) else "Unknown",
                    thread.priority.title(),
                    thread.created_at.strftime("%Y-%m-%d"),
                    thread.last_activity_at.strftime("%Y-%m-%d") if thread.last_activity_at else "",
                    f"{resolution_time:.1f}",  # Hours
                    "Within SLA" if resolution_time <= 24 else "Exceeded SLA",
                    len(thread.comments) if hasattr(thread, 'comments') else 0
                ]

        if export_format.lower() == "csv":
            exporter = CSVExporter(headers, user_locale)
            return StreamingResponse(
                exporter.generate_csv(data_generator()),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=turnaround_report_{datetime.now().strftime('%Y%m%d')}.csv"}
            )
        elif export_format.lower() == "xlsx":
            return await self._handle_xlsx_export(
                headers, data_generator(), f"turnaround_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
        else:
            raise HTTPException(status_code=400, detail="Unsupported export format")

    async def export_bulk_summary(
        self,
        export_format: str,
        filters: ExportFilter,
        tenant_alias: str,
        user_locale: str = "en"
    ) -> StreamingResponse:
        """Export bulk processing summary"""
        headers = self._get_bulk_summary_headers(user_locale)

        def data_generator():
            # Query bulk jobs
            query = select(BulkJob).where(BulkJob.tenant_alias == tenant_alias)

            if filters.start_date:
                query = query.where(BulkJob.created_at >= filters.start_date)
            if filters.end_date:
                query = query.where(BulkJob.created_at <= filters.end_date)
            if filters.status:
                query = query.where(BulkJob.status == filters.status)

            query = query.order_by(desc(BulkJob.created_at))

            result = self.db.execute(query)
            jobs = result.scalars().all()

            for job in jobs:
                # Calculate metrics
                success_rate = 0
                if job.total_items and job.total_items > 0:
                    failed_items = job.failed_items or 0
                    success_rate = ((job.total_items - failed_items) / job.total_items) * 100

                duration = ""
                if job.started_at and job.completed_at:
                    duration_delta = job.completed_at - job.started_at
                    duration = f"{duration_delta.total_seconds() / 60:.1f} min"

                yield [
                    job.id,
                    job.job_type.title(),
                    job.status.title(),
                    job.total_items or 0,
                    (job.total_items or 0) - (job.failed_items or 0),  # Success count
                    job.failed_items or 0,
                    f"{success_rate:.1f}%",
                    duration,
                    job.created_at.strftime("%Y-%m-%d %H:%M"),
                    job.retry_count or 0
                ]

        if export_format.lower() == "csv":
            exporter = CSVExporter(headers, user_locale)
            return StreamingResponse(
                exporter.generate_csv(data_generator()),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=bulk_summary_{datetime.now().strftime('%Y%m%d')}.csv"}
            )
        elif export_format.lower() == "xlsx":
            return await self._handle_xlsx_export(
                headers, data_generator(), f"bulk_summary_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
        else:
            raise HTTPException(status_code=400, detail="Unsupported export format")

    async def _handle_xlsx_export(self, headers: List[str], data_generator, filename: str) -> StreamingResponse:
        """Handle XLSX export with potential S3 storage for large files"""
        try:
            exporter = XLSXExporter(headers)
            xlsx_data = exporter.create_workbook(data_generator)

            # If file is large (>10MB), store in S3 and return signed URL
            if len(xlsx_data) > 10 * 1024 * 1024 and self.s3_client:
                return await self._store_large_file_s3(xlsx_data, filename)

            # Return directly for smaller files
            return StreamingResponse(
                io.BytesIO(xlsx_data),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )

        except Exception as e:
            logger.error(f"XLSX export failed: {str(e)}")
            raise HTTPException(status_code=500, detail="Export generation failed")

    async def _store_large_file_s3(self, data: bytes, filename: str) -> Dict[str, str]:
        """Store large export file in S3 and return signed URL"""
        try:
            bucket_name = settings.S3_BUCKET_NAME
            key = f"exports/{datetime.now().strftime('%Y/%m/%d')}/{filename}"

            # Upload to S3
            self.s3_client.put_object(
                Bucket=bucket_name,
                Key=key,
                Body=data,
                ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ServerSideEncryption="AES256"
            )

            # Generate signed URL (valid for 1 hour)
            signed_url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': bucket_name, 'Key': key},
                ExpiresIn=3600
            )

            logger.info(f"Large export file stored in S3: {key}")
            return {
                "download_url": signed_url,
                "expires_in": 3600,
                "file_size": len(data),
                "filename": filename
            }

        except Exception as e:
            logger.error(f"S3 upload failed: {str(e)}")
            raise HTTPException(status_code=500, detail="File storage failed")

    def _get_portfolio_headers(self, locale: str) -> List[str]:
        """Get localized headers for portfolio report"""
        if locale == "es":
            return ["ID", "Número LC", "Corredor", "Exportador", "Estado", "Prioridad", "Fecha Creación", "Comentarios", "Estado Final"]
        elif locale == "fr":
            return ["ID", "Numéro LC", "Corridor", "Exportateur", "Statut", "Priorité", "Date Création", "Commentaires", "Statut Final"]
        else:  # Default English
            return ["ID", "LC Number", "Corridor", "Exporter", "Status", "Priority", "Created Date", "Comments", "Final Status"]

    def _get_discrepancies_headers(self, locale: str) -> List[str]:
        """Get localized headers for discrepancies report"""
        if locale == "es":
            return ["Fecha", "Tipo Discrepancia", "Conteo", "Severidad", "Tasa Resolución"]
        elif locale == "fr":
            return ["Date", "Type Discordance", "Nombre", "Sévérité", "Taux Résolution"]
        else:  # Default English
            return ["Date", "Discrepancy Type", "Count", "Severity", "Resolution Rate"]

    def _get_turnaround_headers(self, locale: str) -> List[str]:
        """Get localized headers for turnaround report"""
        if locale == "es":
            return ["ID Exportador", "Nombre", "Prioridad", "Fecha Inicio", "Fecha Resolución", "Tiempo (Horas)", "SLA", "Interacciones"]
        elif locale == "fr":
            return ["ID Exportateur", "Nom", "Priorité", "Date Début", "Date Résolution", "Temps (Heures)", "SLA", "Interactions"]
        else:  # Default English
            return ["Exporter ID", "Name", "Priority", "Start Date", "Resolution Date", "Time (Hours)", "SLA Status", "Interactions"]

    def _get_bulk_summary_headers(self, locale: str) -> List[str]:
        """Get localized headers for bulk summary report"""
        if locale == "es":
            return ["ID Trabajo", "Tipo", "Estado", "Total Items", "Éxitos", "Fallos", "Tasa Éxito", "Duración", "Creado", "Reintentos"]
        elif locale == "fr":
            return ["ID Travail", "Type", "Statut", "Total Articles", "Succès", "Échecs", "Taux Succès", "Durée", "Créé", "Tentatives"]
        else:  # Default English
            return ["Job ID", "Type", "Status", "Total Items", "Success", "Failures", "Success Rate", "Duration", "Created", "Retries"]

    def _extract_discrepancy_type(self, title: str) -> str:
        """Extract discrepancy type from thread title (mock implementation)"""
        if "document" in title.lower():
            return "Document Missing"
        elif "amount" in title.lower():
            return "Amount Mismatch"
        elif "date" in title.lower():
            return "Date Discrepancy"
        elif "signature" in title.lower():
            return "Signature Issue"
        else:
            return "Other"

    def _calculate_resolution_time(self, thread) -> float:
        """Calculate resolution time in hours (mock implementation)"""
        if thread.last_activity_at and thread.created_at:
            delta = thread.last_activity_at - thread.created_at
            return delta.total_seconds() / 3600
        else:
            # Mock data for threads without resolution time
            import random
            return random.uniform(2, 48)  # 2-48 hours